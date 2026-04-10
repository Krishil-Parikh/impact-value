"""
Generate the client-facing IoT Assessment data collection Excel template.
Output: frontend/public/IoT_Assessment_Template.xlsx

Layout per data sheet:
  Col A (idx 0) : Row # / section
  Col B (idx 1) : Question label
  Col C (idx 2) : YOUR ANSWER  ← yellow input column
  Col D (idx 3) : Notes / units
  Col E (idx 4) : field_key    ← hidden column, used by parser
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = os.path.join(
    os.path.dirname(__file__),
    "frontend", "public", "IoT_Assessment_Template.xlsx"
)

# ── Colour palette ────────────────────────────────────────────────────────────
C_NAVY   = "1F4E79"
C_BLUE   = "2E74B5"
C_LBLUE  = "D6E4F0"
C_YELLOW = "FFF2CC"   # input cells
C_GREEN  = "E2EFDA"   # barrier section header
C_ORANGE = "FCE4D6"   # cost/kpi section header
C_GRAY   = "F2F2F2"
C_WHITE  = "FFFFFF"
C_RED    = "C00000"

# ── Reusable styles ───────────────────────────────────────────────────────────
def title_style():
    return Font(name="Calibri", size=14, bold=True, color=C_WHITE)

def header_style():
    return Font(name="Calibri", size=10, bold=True, color=C_WHITE)

def label_style(bold=False):
    return Font(name="Calibri", size=10, bold=bold, color="000000")

def key_style():
    return Font(name="Calibri", size=9, color="AAAAAA", italic=True)

def input_style():
    return Font(name="Calibri", size=10, color=C_RED, bold=True)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Helpers ───────────────────────────────────────────────────────────────────

def set_col_widths(ws, widths):
    """widths: dict of letter -> width"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def write_title(ws, text, end_col="E", row=1):
    ws.row_dimensions[row].height = 28
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_style()
    cell.fill = fill(C_NAVY)
    cell.alignment = center()
    ws.merge_cells(f"A{row}:{end_col}{row}")

def write_col_headers(ws, row=2):
    ws.row_dimensions[row].height = 20
    headers = ["#", "Question / Field", "Your Answer", "Notes / Guidance", "field_key"]
    colors  = [C_BLUE, C_BLUE, C_BLUE, C_BLUE, C_BLUE]
    for c, (h, col_fill) in enumerate(zip(headers, colors), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_style()
        cell.fill = fill(col_fill)
        cell.alignment = center()
        cell.border = thin_border()

def section_header(ws, row, text, bg=C_GREEN):
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=10, bold=True, color=C_NAVY)
    cell.fill = fill(bg)
    cell.alignment = left()
    cell.border = thin_border()

def data_row(ws, row, num, label, note, key, dv=None):
    """Write one data row. Returns the input cell."""
    ws.row_dimensions[row].height = 18

    # A: number
    a = ws.cell(row=row, column=1, value=num)
    a.font = label_style()
    a.fill = fill(C_GRAY)
    a.alignment = center()
    a.border = thin_border()

    # B: label
    b = ws.cell(row=row, column=2, value=label)
    b.font = label_style()
    b.fill = fill(C_WHITE)
    b.alignment = left()
    b.border = thin_border()

    # C: input (yellow)
    c = ws.cell(row=row, column=3, value=None)
    c.font = input_style()
    c.fill = fill(C_YELLOW)
    c.alignment = center()
    c.border = thin_border()
    if dv:
        dv.add(c)

    # D: notes
    d = ws.cell(row=row, column=4, value=note)
    d.font = Font(name="Calibri", size=9, color="555555", italic=True)
    d.fill = fill(C_WHITE)
    d.alignment = left()
    d.border = thin_border()

    # E: hidden key
    e = ws.cell(row=row, column=5, value=key)
    e.font = key_style()
    e.fill = fill(C_WHITE)
    e.alignment = left()
    e.border = thin_border()

    return c


# ── Barrier data ──────────────────────────────────────────────────────────────

BARRIERS = [
    ("1",  "Lack of Training for Workers and Managers", [
        ("num_training_programs",         "Number of training programs offered per year",                    "Enter count (if >5, enter 5)"),
        ("pct_employees_trained",         "% of employees trained in digital technologies",                  "Enter percentage, e.g. 30 for 30%"),
        ("pct_budget_training_of_payroll","Training budget as % of total payroll",                           "Enter percentage, e.g. 2.5 for 2.5%"),
    ]),
    ("2",  "Resistance to Change", [
        ("employee_turnover_rate_pct",    "Employee turnover rate following IoT initiatives (%)",            "Enter percentage, e.g. 15 for 15%"),
        ("pct_employees_resisting",       "% of employees expressing resistance to new technologies",        "Enter percentage, e.g. 20 for 20%"),
        ("num_feedback_sessions",         "Number of feedback/concern sessions held per year",               "Enter count (12/yr is healthy)"),
    ]),
    ("3",  "Lack of Digital Culture and Training", [
        ("num_digital_skills_workshops",          "Number of digital skills workshops attended by employees",    "Enter count (5/yr is healthy)"),
        ("pct_comfortable_digital_tools",         "% of employees comfortable using digital tools",             "Enter percentage, e.g. 30 for 30%"),
        ("adoption_rate_new_digital_tools_pct",   "Adoption rate of new digital tools across the organisation (%)","Enter percentage, e.g. 25 for 25%"),
    ]),
    ("4",  "Higher Investment in Employees' Training", [
        ("pct_training_expenditure_of_op_costs",  "Training expenditure as % of total operational costs",       "Enter percentage, e.g. 1.25 for 1.25%"),
        ("avg_training_hours_per_employee",        "Average training hours per employee per year",               "Enter hours (40 hrs is the norm)"),
        ("roi_training_programs_pct",              "ROI on training programs (%)",                               "Enter percentage, e.g. 5 for 5%"),
    ]),
    ("5",  "Lack of Knowledge Management Systems", [
        ("num_knowledge_sharing_sessions",         "Number of knowledge-sharing sessions conducted internally per year","Enter count (24/yr is healthy)"),
        ("pct_employees_access_kms",               "% of employees with access to a centralised knowledge system","Enter percentage, e.g. 20 for 20%"),
        ("freq_updates_kms",                       "Frequency of updates to the knowledge management system",   "Select from dropdown", "kms"),
    ]),
    ("6",  "Regulatory Compliance Issues", [
        ("num_non_compliance_incidents",            "Number of non-compliance incidents or penalties per year",  "Enter count (5+/yr = crisis)"),
        ("pct_projects_delayed_regulatory",         "% of IoT projects delayed due to regulatory challenges",   "Enter percentage, e.g. 40 for 40%"),
        ("time_to_achieve_compliance_days",         "Time to achieve compliance with new IoT regulations (days)","Enter days (180 days = crisis)"),
    ]),
    ("7",  "Lack of Standards and Reference Architecture", [
        ("num_industry_standards_adopted",              "Number of industry standards adopted by the organisation","Enter count (10 is a healthy target)"),
        ("pct_iot_devices_conforming",                  "% of IoT devices conforming to reference architectures", "Enter percentage, e.g. 50 for 50%"),
        ("pct_projects_delayed_standardized_solutions", "% of projects delayed due to lack of standardised solutions","Enter percentage, e.g. 50 for 50%"),
    ]),
    ("8",  "Lack of Internet Coverage and IT Facilities", [
        ("pct_internet_coverage",                       "% of work locations with stable internet connectivity",  "Enter percentage, e.g. 70 for 70%"),
        ("avg_internet_speed_mbps",                     "Average internet speed across different locations (Mbps)","Enter speed (50 Mbps is good)"),
        ("freq_it_infrastructure_outages_per_month",    "Frequency of IT infrastructure outages per month",       "Enter count (8+/month = crisis)"),
    ]),
    ("9",  "Limited Access to Funding and Credit", [
        ("pct_loan_approved",                               "% of loan applications approved for technology implementation","Enter percentage, e.g. 80 for 80%"),
        ("pct_projects_delayed_lack_funding",               "% of projects delayed or cancelled due to lack of funding","Enter percentage, e.g. 30 for 30%"),
        ("ratio_external_funding_total_project_costs_pct",  "External funding (grants/loans) as % of total project costs","Enter percentage, e.g. 15 for 15%"),
    ]),
    ("10", "Future Viability & Profitability", [
        ("yoy_revenue_growth_iot_pct",           "Year-over-year revenue growth attributed to IoT (%)",         "Enter percentage, e.g. 2 for 2%"),
        ("profit_margin_improvement_iot_pct",    "Profit margin improvement following IoT implementation (%)",  "Enter percentage, e.g. 0.5 for 0.5%"),
        ("num_new_revenue_streams_iot",          "Number of new revenue streams generated through IoT",         "Enter count (3 = successful target)"),
    ]),
    ("11", "Dependency on External Vendors", [
        ("pct_critical_operations_reliant_vendor",      "% of critical operations reliant on third-party vendors","Enter percentage, e.g. 80 for 80%"),
        ("num_vendor_delays_disruptions_per_year",      "Number of vendor-related delays or disruptions per year","Enter count"),
        ("cost_vendor_contracts_pct_op_expenses",       "Cost of vendor contracts as % of total operating expenses","Enter percentage, e.g. 35 for 35%"),
    ]),
    ("12", "High Implementation Cost", [
        ("pct_it_budget_allocated_iot",                 "% of total IT budget allocated to IoT implementation",  "Enter percentage, e.g. 50 for 50%"),
        ("annual_maintenance_costs_pct_op_costs",       "Annual maintenance & support costs as % of operational costs","Enter percentage (5%+ = crisis for SME)"),
        ("integration_costs_pct_total_project_cost",    "Integration costs as % of total project cost",          "Enter percentage, e.g. 40 for 40%"),
    ]),
    ("13", "Compliance with Sector-Specific Regulations", [
        ("num_regulatory_violations_penalties",             "Number of regulatory violations or penalties per year","Enter count (4+/yr = crisis)"),
        ("pct_compliance_audits_passed_without_issues",     "% of compliance audits passed without issues",        "Enter percentage, e.g. 60 for 60%"),
        ("freq_updates_internal_policies",                  "Frequency of updates to internal policies/procedures","Select from dropdown", "policy"),
    ]),
    ("14", "Lack of Regulations and Standards", [
        ("pct_standards_compliant_iot_devices",             "% of standards-compliant IoT devices used",           "Enter percentage, e.g. 30 for 30%"),
        ("num_industry_specific_guidelines_implemented",    "Number of industry-specific guidelines implemented",  "Enter count (10 = healthy target)"),
    ]),
    ("15", "Customers are Hesitant to Share Data", [
        ("pct_customers_refuse_data",                       "% of customers who refuse to share data (survey-based)","Enter percentage, e.g. 40 for 40%"),
        ("num_customer_complaints_data_sharing",            "Number of customer complaints/concerns on data sharing","Enter count (25+/yr = crisis)"),
        ("pct_customer_contracts_explicit_data_sharing",    "% of customer contracts with explicit data-sharing agreements","Enter percentage, e.g. 50 for 50%"),
    ]),
]

COST_FACTORS = [
    ("aftermarket_services_warranty",           "Aftermarket Services / Warranty",                          "operational"),
    ("depreciation",                            "Depreciation",                                             "operational"),
    ("labour",                                  "Labour",                                                   "operational"),
    ("maintenance_repair",                      "Maintenance & Repair",                                     "operational"),
    ("raw_materials_consumables",               "Raw Materials & Consumables",                              "operational"),
    ("rental_operating_lease",                  "Rental & Operating Lease",                                 "operational"),
    ("research_development",                    "Research & Development (R&D)",                             "strategic"),
    ("selling_general_administrative_expense",  "Selling, General & Administrative Expense",                "operational"),
    ("utilities",                               "Utilities",                                                "operational"),
    ("earnings_before_interest_taxes_ebit",     "Earnings Before Interest & Taxes (EBIT)",                  "financial"),
    ("financing_costs_interest",                "Financing Costs (Interest)",                               "financial"),
    ("taxation_compliance_costs",               "Taxation & Compliance Costs",                              "financial"),
    ("supply_chain_logistics_costs",            "Supply Chain & Logistics Costs",                           "operational"),
    ("technology_digital_infrastructure_costs", "Technology & Digital Infrastructure Costs",                "strategic"),
    ("training_skill_development_costs",        "Training & Skill Development Costs",                       "strategic"),
    ("regulatory_compliance_costs",             "Regulatory & Compliance Costs",                            "financial"),
    ("insurance_costs",                         "Insurance Costs",                                          "financial"),
    ("marketing_customer_acquisition_costs",    "Marketing & Customer Acquisition Costs",                   "strategic"),
    ("environmental_social_responsibility_costs","Environmental & Social Responsibility Costs",             "strategic"),
    ("quality_control_assurance",               "Quality Control & Assurance",                              "operational"),
]

KPI_FACTORS = [
    ("asset_equipment_efficiency",          "Asset & Equipment Efficiency",         "operational"),
    ("utilities_efficiency",                "Utilities Efficiency",                 "operational"),
    ("inventory_efficiency",                "Inventory Efficiency",                 "operational"),
    ("process_quality",                     "Process Quality",                      "quality"),
    ("product_quality",                     "Product Quality",                      "quality"),
    ("safety_security",                     "Safety & Security",                    "quality"),
    ("planning_scheduling_effectiveness",   "Planning & Scheduling Effectiveness",  "operational"),
    ("time_to_market",                      "Time to Market",                       "performance"),
    ("production_flexibility",              "Production Flexibility",               "operational"),
    ("customer_satisfaction",               "Customer Satisfaction",                "customer"),
    ("supply_chain_efficiency",             "Supply Chain Efficiency",              "operational"),
    ("market_share_growth",                 "Market Share Growth",                  "performance"),
    ("employee_productivity",               "Employee Productivity",                "performance"),
    ("return_on_investment_roi",            "Return on Investment (ROI)",           "financial"),
    ("financial_health_and_stability",      "Financial Health & Stability",         "financial"),
    ("talent_retention",                    "Talent Retention",                     "customer"),
    ("customer_retention_rate",             "Customer Retention Rate",              "customer"),
]


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 3, "B": 80})

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "IoT Readiness Assessment — Client Data Collection Template"
    c.font = Font(name="Calibri", size=16, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 36

    lines = [
        ("HOW TO USE THIS TEMPLATE", True),
        ("", False),
        ("1.  This workbook contains 4 data sheets:", False),
        ("       • Company Details   – Basic information about your organisation", False),
        ("       • Barrier Assessment – 15 barrier questions (43 fields total)", False),
        ("       • Cost Factors       – 20 cost categories as a decimal fraction of annual revenue", False),
        ("       • KPI Factors        – 17 KPIs your organisation tracks (enter Yes or No)", False),
        ("", False),
        ("2.  Only fill in the yellow-highlighted cells in column C of each sheet.", False),
        ("       Do NOT modify any other columns — especially the hidden 'field_key' column.", False),
        ("", False),
        ("3.  COST FACTORS — enter values as a decimal fraction of annual revenue.", False),
        ("       Example:  If Labour costs are 13% of your revenue, enter 0.13", False),
        ("       Example:  If Raw Materials are 42% of your revenue, enter 0.42", False),
        ("       All 20 cost factors should ideally sum to ≈ 1.00 (100%).", False),
        ("", False),
        ("4.  BARRIER ASSESSMENT — enter numeric values or select from the dropdown menus.", False),
        ("       Percentages are entered as whole numbers (e.g., enter 30 for 30%).", False),
        ("", False),
        ("5.  KPI FACTORS — select Yes if your organisation tracks that KPI, No if not.", False),
        ("", False),
        ("6.  Save the completed file and upload it using the 'Upload Excel' button", False),
        ("       on the IoT Readiness Assessment portal.", False),
        ("", False),
        ("CONTACT:  If you have questions, please reach out to your assessment coordinator.", False),
    ]

    for i, (text, bold) in enumerate(lines, start=2):
        ws.row_dimensions[i + 1].height = 16
        c = ws.cell(row=i + 1, column=2, value=text)
        c.font = Font(name="Calibri", size=11, bold=bold,
                      color=C_NAVY if bold else "333333")
        c.alignment = left()

    ws.sheet_state = "visible"


def build_company_sheet(wb):
    ws = wb.create_sheet("Company Details")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 6, "B": 45, "C": 28, "D": 35, "E": 0.1})
    ws.column_dimensions["E"].hidden = True

    write_title(ws, "COMPANY DETAILS", row=1)
    write_col_headers(ws, row=2)

    fields = [
        ("company_name",    "Company Name",                       "Enter full company name"),
        ("industry",        "Industry",                           "e.g. Manufacturing, Automotive, Textiles"),
        ("num_employees",   "Number of Employees",                "Enter total headcount (integer)"),
        ("annual_revenue",  "Annual Revenue (in Crores ₹)",       "Enter value in crores"),
    ]

    for i, (key, label, note) in enumerate(fields, start=1):
        data_row(ws, row=i + 2, num=i, label=label, note=note, key=key)

    ws.freeze_panes = "A3"


def build_barrier_sheet(wb):
    ws = wb.create_sheet("Barrier Assessment")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 6, "B": 58, "C": 22, "D": 38, "E": 0.1})
    ws.column_dimensions["E"].hidden = True

    write_title(ws, "BARRIER ASSESSMENT — 15 Barriers to IoT Adoption", row=1)
    write_col_headers(ws, row=2)

    # Data validations
    dv_kms = DataValidation(
        type="list",
        formula1='"Daily,Weekly,Monthly,Quarterly,Bi-Annually,Annually"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid selection",
        error="Please select a valid frequency from the dropdown."
    )
    dv_policy = DataValidation(
        type="list",
        formula1='"Monthly,Quarterly,Bi-Annually,Annually"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid selection",
        error="Please select a valid frequency from the dropdown."
    )
    ws.add_data_validation(dv_kms)
    ws.add_data_validation(dv_policy)

    current_row = 3
    field_num = 1

    for barrier_num, barrier_title, fields in BARRIERS:
        # Barrier section header
        section_header(ws, current_row,
                       f"  BARRIER {barrier_num}:  {barrier_title}", bg=C_LBLUE)
        current_row += 1

        for field_tuple in fields:
            key  = field_tuple[0]
            label = field_tuple[1]
            note  = field_tuple[2]
            dv_type = field_tuple[3] if len(field_tuple) > 3 else None

            if dv_type == "kms":
                dv = dv_kms
            elif dv_type == "policy":
                dv = dv_policy
            else:
                dv = None

            data_row(ws, row=current_row, num=field_num,
                     label=label, note=note, key=key, dv=dv)
            current_row += 1
            field_num += 1

    ws.freeze_panes = "A3"


def build_cost_sheet(wb):
    ws = wb.create_sheet("Cost Factors")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 6, "B": 55, "C": 22, "D": 40, "E": 0.1})
    ws.column_dimensions["E"].hidden = True

    write_title(ws, "COST FACTORS — As a Decimal Fraction of Annual Revenue", row=1)

    # Sub-instruction row
    ws.row_dimensions[2].height = 30
    ws.merge_cells("A2:E2")
    ci = ws["A2"]
    ci.value = (
        "⚠  Enter each value as a decimal fraction of annual revenue.  "
        "Example: if Labour = 13% of revenue → enter 0.13     "
        "All 20 values should ideally total ≈ 1.00"
    )
    ci.font = Font(name="Calibri", size=10, bold=True, color=C_NAVY)
    ci.fill = fill(C_YELLOW)
    ci.alignment = center()
    ci.border = thin_border()

    write_col_headers(ws, row=3)

    CAT_COLORS = {
        "operational": C_WHITE,
        "strategic":   "F0F7FF",
        "financial":   "FFF8F0",
    }
    CAT_LABELS = {
        "operational": "Operational",
        "strategic":   "Strategic Investment",
        "financial":   "Financial",
    }

    prev_cat = None
    row = 4
    for i, (key, label, cat) in enumerate(COST_FACTORS, start=1):
        if cat != prev_cat:
            section_header(ws, row,
                           f"  {CAT_LABELS[cat]} Costs", bg=C_ORANGE)
            row += 1
            prev_cat = cat

        c = data_row(ws, row=row, num=i, label=label,
                     note="Decimal fraction (e.g. 0.13 for 13%)", key=key)
        # Force number format
        c.number_format = "0.0000"
        row += 1

    ws.freeze_panes = "A4"


def build_kpi_sheet(wb):
    ws = wb.create_sheet("KPI Factors")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 6, "B": 50, "C": 18, "D": 40, "E": 0.1})
    ws.column_dimensions["E"].hidden = True

    write_title(ws, "KPI FACTORS — Which KPIs Does Your Organisation Track?", row=1)

    ws.row_dimensions[2].height = 28
    ws.merge_cells("A2:E2")
    ci = ws["A2"]
    ci.value = (
        "Select  Yes  if your organisation actively tracks and considers this KPI important.  "
        "Select  No  if it is not currently measured or relevant."
    )
    ci.font = Font(name="Calibri", size=10, bold=True, color=C_NAVY)
    ci.fill = fill(C_YELLOW)
    ci.alignment = center()
    ci.border = thin_border()

    write_col_headers(ws, row=3)

    # Yes/No validation
    dv_yn = DataValidation(
        type="list",
        formula1='"Yes,No"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid",
        error="Please select Yes or No."
    )
    ws.add_data_validation(dv_yn)

    CAT_LABELS = {
        "operational": "Operational Excellence",
        "quality":     "Quality & Safety",
        "performance": "Performance Metrics",
        "financial":   "Financial Performance",
        "customer":    "Customer & People",
    }

    prev_cat = None
    row = 4
    for i, (key, label, cat) in enumerate(KPI_FACTORS, start=1):
        if cat != prev_cat:
            section_header(ws, row, f"  {CAT_LABELS[cat]}", bg=C_ORANGE)
            row += 1
            prev_cat = cat

        c = data_row(ws, row=row, num=i, label=label,
                     note="Select: Yes or No", key=key, dv=dv_yn)
        row += 1

    ws.freeze_panes = "A4"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    build_instructions(wb)
    build_company_sheet(wb)
    build_barrier_sheet(wb)
    build_cost_sheet(wb)
    build_kpi_sheet(wb)

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"Template saved: {OUTPUT}")


if __name__ == "__main__":
    main()
