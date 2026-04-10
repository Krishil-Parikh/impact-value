"""
Create a filled IoT Assessment Template for Plast Fab Works Pvt. Ltd.
Uses the same template structure as generate_client_template.py so that
the "Upload Excel" button in the frontend can parse it correctly.
"""
import os
import sys
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = os.path.join(
    os.path.dirname(__file__),
    "Plast_Fab_Works_Assessment_Template.xlsx"
)

# ── Colours ───────────────────────────────────────────────────────────────────
C_NAVY   = "1F4E79"
C_BLUE   = "2E74B5"
C_LBLUE  = "D6E4F0"
C_YELLOW = "FFF2CC"
C_GREEN  = "E2EFDA"
C_ORANGE = "FCE4D6"
C_GRAY   = "F2F2F2"
C_WHITE  = "FFFFFF"
C_RED    = "C00000"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def write_title(ws, text, end_col="E", row=1):
    ws.row_dimensions[row].height = 28
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=14, bold=True, color=C_WHITE)
    cell.fill = fill(C_NAVY)
    cell.alignment = center()
    ws.merge_cells(f"A{row}:{end_col}{row}")

def write_col_headers(ws, row=2):
    ws.row_dimensions[row].height = 20
    headers = ["#", "Question / Field", "Your Answer", "Notes / Guidance", "field_key"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
        cell.fill = fill(C_BLUE)
        cell.alignment = center()
        cell.border = thin_border()

def section_header(ws, row, text, bg=C_GREEN):
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=10, bold=True, color=C_NAVY)
    cell.fill = fill(bg)
    cell.alignment = left()
    cell.border = thin_border()

def data_row(ws, row, num, label, note, key, value=None, dv=None):
    ws.row_dimensions[row].height = 18
    a = ws.cell(row=row, column=1, value=num)
    a.font = Font(name="Calibri", size=10); a.fill = fill(C_GRAY)
    a.alignment = center(); a.border = thin_border()

    b = ws.cell(row=row, column=2, value=label)
    b.font = Font(name="Calibri", size=10); b.fill = fill(C_WHITE)
    b.alignment = left(); b.border = thin_border()

    c = ws.cell(row=row, column=3, value=value)
    c.font = Font(name="Calibri", size=10, color=C_RED, bold=True)
    c.fill = fill(C_YELLOW); c.alignment = center(); c.border = thin_border()
    if dv:
        dv.add(c)

    d = ws.cell(row=row, column=4, value=note)
    d.font = Font(name="Calibri", size=9, color="555555", italic=True)
    d.fill = fill(C_WHITE); d.alignment = left(); d.border = thin_border()

    e = ws.cell(row=row, column=5, value=key)
    e.font = Font(name="Calibri", size=9, color="AAAAAA", italic=True)
    e.fill = fill(C_WHITE); e.alignment = left(); e.border = thin_border()

    return c


# ── Plast Fab Works data ──────────────────────────────────────────────────────

COMPANY = {
    "company_name":   "Plast Fab Works Pvt. Ltd.",
    "industry":       "Plastic injection molding",
    "num_employees":  30,
    "annual_revenue": 3,
}

# Reverse-engineered from Plastfabworks survey using backend barrier_service.py formulas
BARRIERS_DATA = {
    "barrier1":  {
        "num_training_programs": 2,
        "pct_employees_trained": 10,
        "pct_budget_training_of_payroll": 1.667,
    },
    "barrier2":  {
        "employee_turnover_rate_pct": 0,
        "pct_employees_resisting": 40,
        "num_feedback_sessions": 1,
    },
    "barrier3":  {
        "num_digital_skills_workshops": 3,
        "pct_comfortable_digital_tools": 30,
        "adoption_rate_new_digital_tools_pct": 25,
    },
    "barrier4":  {
        "pct_training_expenditure_of_op_costs": 1.25,
        "avg_training_hours_per_employee": 30,
        "roi_training_programs_pct": 5,
    },
    "barrier5":  {
        "num_knowledge_sharing_sessions": 2,
        "pct_employees_access_kms": 20,
        "freq_updates_kms": "Quarterly",
    },
    "barrier6":  {
        "num_non_compliance_incidents": 0,
        "pct_projects_delayed_regulatory": 40,
        "time_to_achieve_compliance_days": 90,
    },
    "barrier7":  {
        "num_industry_standards_adopted": 7,
        "pct_iot_devices_conforming": 0,
        "pct_projects_delayed_standardized_solutions": 50,
    },
    "barrier8":  {
        "pct_internet_coverage": 70,
        "avg_internet_speed_mbps": 100,
        "freq_it_infrastructure_outages_per_month": 2,
    },
    "barrier9":  {
        "pct_loan_approved": 0,
        "pct_projects_delayed_lack_funding": 30,
        "ratio_external_funding_total_project_costs_pct": 15,
    },
    "barrier10": {
        "yoy_revenue_growth_iot_pct": 0.3,
        "profit_margin_improvement_iot_pct": 0.5,
        "num_new_revenue_streams_iot": 0,
    },
    "barrier11": {
        "pct_critical_operations_reliant_vendor": 80,
        "num_vendor_delays_disruptions_per_year": 15,
        "cost_vendor_contracts_pct_op_expenses": 35,
    },
    "barrier12": {
        "pct_it_budget_allocated_iot": 50,
        "annual_maintenance_costs_pct_op_costs": 0.05,
        "integration_costs_pct_total_project_cost": 40,
    },
    "barrier13": {
        "num_regulatory_violations_penalties": 0,
        "pct_compliance_audits_passed_without_issues": 60,
        "freq_updates_internal_policies": "Annually",
    },
    "barrier14": {
        "pct_standards_compliant_iot_devices": 30,
        "num_industry_specific_guidelines_implemented": 0,
    },
    "barrier15": {
        "pct_customers_refuse_data": 40,
        "num_customer_complaints_data_sharing": 0,
        "pct_customer_contracts_explicit_data_sharing": 20,
    },
}

COST_DATA = {
    "aftermarket_services_warranty": 0,
    "depreciation": 0.02,
    "labour": 0.04,
    "maintenance_repair": 0.004,
    "raw_materials_consumables": 0.024,
    "rental_operating_lease": 0,
    "research_development": 0,
    "selling_general_administrative_expense": 0.002,
    "utilities": 0.032,
    "earnings_before_interest_taxes_ebit": 0.6,
    "financing_costs_interest": 0,
    "taxation_compliance_costs": 0.1,
    "supply_chain_logistics_costs": 0.008,
    "technology_digital_infrastructure_costs": 0.006,
    "training_skill_development_costs": 0.002,
    "regulatory_compliance_costs": 0.0028,
    "insurance_costs": 0.0096,
    "marketing_customer_acquisition_costs": 0.00048,
    "environmental_social_responsibility_costs": 0,
    "quality_control_assurance": 0.0008,
}

KPI_DATA = {
    "asset_equipment_efficiency": "Yes",
    "utilities_efficiency": "No",
    "inventory_efficiency": "No",
    "process_quality": "Yes",
    "product_quality": "No",
    "safety_security": "No",
    "planning_scheduling_effectiveness": "Yes",
    "time_to_market": "No",
    "production_flexibility": "Yes",
    "customer_satisfaction": "No",
    "supply_chain_efficiency": "No",
    "market_share_growth": "Yes",
    "employee_productivity": "Yes",
    "return_on_investment_roi": "No",
    "financial_health_and_stability": "No",
    "talent_retention": "No",
    "customer_retention_rate": "No",
}

# ── Barrier definitions (same as generate_client_template.py) ─────────────────

BARRIERS = [
    ("1",  "Lack of Training for Workers and Managers", [
        ("num_training_programs",          "Number of training programs offered per year",             "Enter count (if >5, enter 5)"),
        ("pct_employees_trained",          "% of employees trained in digital technologies",           "Enter percentage, e.g. 30 for 30%"),
        ("pct_budget_training_of_payroll", "Training budget as % of total payroll",                    "Enter percentage, e.g. 2.5 for 2.5%"),
    ]),
    ("2",  "Resistance to Change", [
        ("employee_turnover_rate_pct",  "Employee turnover rate following IoT initiatives (%)",       "Enter percentage, e.g. 15"),
        ("pct_employees_resisting",     "% of employees expressing resistance",                       "Enter percentage, e.g. 20"),
        ("num_feedback_sessions",       "Number of feedback sessions held per year",                  "Enter count"),
    ]),
    ("3",  "Lack of Digital Culture and Training", [
        ("num_digital_skills_workshops",        "Number of digital skills workshops attended",        "Enter count"),
        ("pct_comfortable_digital_tools",       "% of employees comfortable using digital tools",     "Enter percentage"),
        ("adoption_rate_new_digital_tools_pct", "Adoption rate of new digital tools (%)",             "Enter percentage"),
    ]),
    ("4",  "Higher Investment in Employees' Training", [
        ("pct_training_expenditure_of_op_costs", "Training expenditure as % of operational costs",   "Enter percentage"),
        ("avg_training_hours_per_employee",       "Average training hours per employee per year",     "Enter hours"),
        ("roi_training_programs_pct",             "ROI on training programs (%)",                     "Enter percentage"),
    ]),
    ("5",  "Lack of Knowledge Management Systems", [
        ("num_knowledge_sharing_sessions", "Knowledge-sharing sessions per year",                     "Enter count"),
        ("pct_employees_access_kms",       "% of employees with access to knowledge system",          "Enter percentage"),
        ("freq_updates_kms",               "Frequency of KMS updates",                               "Select from dropdown", "kms"),
    ]),
    ("6",  "Regulatory Compliance Issues", [
        ("num_non_compliance_incidents",     "Non-compliance incidents per year",                     "Enter count"),
        ("pct_projects_delayed_regulatory",  "% of IoT projects delayed due to regulatory issues",   "Enter percentage"),
        ("time_to_achieve_compliance_days",  "Time to achieve compliance (days)",                     "Enter days"),
    ]),
    ("7",  "Lack of Standards and Reference Architecture", [
        ("num_industry_standards_adopted",              "Industry standards adopted",                 "Enter count"),
        ("pct_iot_devices_conforming",                  "% of IoT devices conforming to standards",  "Enter percentage"),
        ("pct_projects_delayed_standardized_solutions", "% of projects delayed due to no standards", "Enter percentage"),
    ]),
    ("8",  "Lack of Internet Coverage and IT Facilities", [
        ("pct_internet_coverage",                    "% of locations with stable internet",           "Enter percentage"),
        ("avg_internet_speed_mbps",                  "Average internet speed (Mbps)",                 "Enter speed"),
        ("freq_it_infrastructure_outages_per_month", "IT outages per month",                          "Enter count"),
    ]),
    ("9",  "Limited Access to Funding and Credit", [
        ("pct_loan_approved",                            "% of loan applications approved",           "Enter percentage"),
        ("pct_projects_delayed_lack_funding",            "% of projects delayed due to lack of funds","Enter percentage"),
        ("ratio_external_funding_total_project_costs_pct","External funding as % of project cost",   "Enter percentage"),
    ]),
    ("10", "Future Viability & Profitability", [
        ("yoy_revenue_growth_iot_pct",        "YoY revenue growth attributed to IoT (%)",            "Enter percentage"),
        ("profit_margin_improvement_iot_pct", "Profit margin improvement after IoT (%)",              "Enter percentage"),
        ("num_new_revenue_streams_iot",        "New revenue streams from IoT",                        "Enter count"),
    ]),
    ("11", "Dependency on External Vendors", [
        ("pct_critical_operations_reliant_vendor",  "% of critical operations reliant on vendors",   "Enter percentage"),
        ("num_vendor_delays_disruptions_per_year",  "Vendor delays/disruptions per year",             "Enter count"),
        ("cost_vendor_contracts_pct_op_expenses",   "Vendor contract cost as % of operating expenses","Enter percentage"),
    ]),
    ("12", "High Implementation Cost", [
        ("pct_it_budget_allocated_iot",             "% of IT budget allocated to IoT",               "Enter percentage"),
        ("annual_maintenance_costs_pct_op_costs",   "Annual maintenance cost as % of op costs",      "Enter percentage"),
        ("integration_costs_pct_total_project_cost","Integration costs as % of project cost",        "Enter percentage"),
    ]),
    ("13", "Compliance with Sector-Specific Regulations", [
        ("num_regulatory_violations_penalties",         "Regulatory violations per year",             "Enter count"),
        ("pct_compliance_audits_passed_without_issues", "% of compliance audits passed clean",        "Enter percentage"),
        ("freq_updates_internal_policies",              "Frequency of internal policy updates",       "Select from dropdown", "policy"),
    ]),
    ("14", "Lack of Regulations and Standards", [
        ("pct_standards_compliant_iot_devices",          "% of standards-compliant IoT devices",     "Enter percentage"),
        ("num_industry_specific_guidelines_implemented", "Industry-specific guidelines implemented",  "Enter count"),
    ]),
    ("15", "Customers are Hesitant to Share Data", [
        ("pct_customers_refuse_data",                    "% of customers refusing to share data",    "Enter percentage"),
        ("num_customer_complaints_data_sharing",         "Customer complaints on data sharing",       "Enter count"),
        ("pct_customer_contracts_explicit_data_sharing", "% of contracts with data-sharing clauses", "Enter percentage"),
    ]),
]

COST_FACTORS = [
    ("aftermarket_services_warranty",           "Aftermarket Services / Warranty",                "operational"),
    ("depreciation",                            "Depreciation",                                   "operational"),
    ("labour",                                  "Labour",                                         "operational"),
    ("maintenance_repair",                      "Maintenance & Repair",                           "operational"),
    ("raw_materials_consumables",               "Raw Materials & Consumables",                    "operational"),
    ("rental_operating_lease",                  "Rental & Operating Lease",                       "operational"),
    ("research_development",                    "Research & Development (R&D)",                   "strategic"),
    ("selling_general_administrative_expense",  "Selling, General & Administrative Expense",      "operational"),
    ("utilities",                               "Utilities",                                      "operational"),
    ("earnings_before_interest_taxes_ebit",     "Earnings Before Interest & Taxes (EBIT)",        "financial"),
    ("financing_costs_interest",                "Financing Costs (Interest)",                     "financial"),
    ("taxation_compliance_costs",               "Taxation & Compliance Costs",                    "financial"),
    ("supply_chain_logistics_costs",            "Supply Chain & Logistics Costs",                 "operational"),
    ("technology_digital_infrastructure_costs", "Technology & Digital Infrastructure Costs",      "strategic"),
    ("training_skill_development_costs",        "Training & Skill Development Costs",             "strategic"),
    ("regulatory_compliance_costs",             "Regulatory & Compliance Costs",                  "financial"),
    ("insurance_costs",                         "Insurance Costs",                                "financial"),
    ("marketing_customer_acquisition_costs",    "Marketing & Customer Acquisition Costs",         "strategic"),
    ("environmental_social_responsibility_costs","Environmental & Social Responsibility Costs",   "strategic"),
    ("quality_control_assurance",               "Quality Control & Assurance",                    "operational"),
]

KPI_FACTORS = [
    ("asset_equipment_efficiency",          "Asset & Equipment Efficiency",        "operational"),
    ("utilities_efficiency",                "Utilities Efficiency",                "operational"),
    ("inventory_efficiency",                "Inventory Efficiency",                "operational"),
    ("process_quality",                     "Process Quality",                     "quality"),
    ("product_quality",                     "Product Quality",                     "quality"),
    ("safety_security",                     "Safety & Security",                   "quality"),
    ("planning_scheduling_effectiveness",   "Planning & Scheduling Effectiveness", "operational"),
    ("time_to_market",                      "Time to Market",                      "performance"),
    ("production_flexibility",              "Production Flexibility",              "operational"),
    ("customer_satisfaction",               "Customer Satisfaction",               "customer"),
    ("supply_chain_efficiency",             "Supply Chain Efficiency",             "operational"),
    ("market_share_growth",                 "Market Share Growth",                 "performance"),
    ("employee_productivity",               "Employee Productivity",               "performance"),
    ("return_on_investment_roi",            "Return on Investment (ROI)",          "financial"),
    ("financial_health_and_stability",      "Financial Health & Stability",        "financial"),
    ("talent_retention",                    "Talent Retention",                    "customer"),
    ("customer_retention_rate",             "Customer Retention Rate",             "customer"),
]


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_company_sheet(wb):
    ws = wb.create_sheet("Company Details")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 6, "B": 45, "C": 28, "D": 35, "E": 0.1})
    ws.column_dimensions["E"].hidden = True
    write_title(ws, "COMPANY DETAILS — Plast Fab Works Pvt. Ltd.", row=1)
    write_col_headers(ws, row=2)
    fields = [
        ("company_name",   "Company Name",                 "Enter full company name"),
        ("industry",       "Industry",                     "e.g. Manufacturing, Automotive"),
        ("num_employees",  "Number of Employees",          "Enter total headcount"),
        ("annual_revenue", "Annual Revenue (in Crores Rs)","Enter value in crores"),
    ]
    for i, (key, label, note) in enumerate(fields, start=1):
        data_row(ws, row=i+2, num=i, label=label, note=note, key=key,
                 value=COMPANY[key])
    ws.freeze_panes = "A3"


def build_barrier_sheet(wb):
    ws = wb.create_sheet("Barrier Assessment")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 6, "B": 58, "C": 22, "D": 38, "E": 0.1})
    ws.column_dimensions["E"].hidden = True
    write_title(ws, "BARRIER ASSESSMENT — Plast Fab Works Pvt. Ltd.", row=1)
    write_col_headers(ws, row=2)

    dv_kms = DataValidation(type="list", formula1='"Daily,Weekly,Monthly,Quarterly,Bi-Annually,Annually"',
                            showDropDown=False, showErrorMessage=True)
    dv_policy = DataValidation(type="list", formula1='"Monthly,Quarterly,Bi-Annually,Annually"',
                               showDropDown=False, showErrorMessage=True)
    ws.add_data_validation(dv_kms)
    ws.add_data_validation(dv_policy)

    current_row = 3
    field_num = 1

    for barrier_num, barrier_title, fields in BARRIERS:
        b_key = f"barrier{barrier_num}"
        b_data = BARRIERS_DATA.get(b_key, {})

        section_header(ws, current_row, f"  BARRIER {barrier_num}:  {barrier_title}", bg=C_LBLUE)
        current_row += 1

        for field_tuple in fields:
            key   = field_tuple[0]
            label = field_tuple[1]
            note  = field_tuple[2]
            dv_type = field_tuple[3] if len(field_tuple) > 3 else None
            value = b_data.get(key, None)

            dv = dv_kms if dv_type == "kms" else (dv_policy if dv_type == "policy" else None)
            data_row(ws, row=current_row, num=field_num, label=label, note=note,
                     key=key, value=value, dv=dv)
            current_row += 1
            field_num += 1

    ws.freeze_panes = "A3"


def build_cost_sheet(wb):
    ws = wb.create_sheet("Cost Factors")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 6, "B": 55, "C": 22, "D": 40, "E": 0.1})
    ws.column_dimensions["E"].hidden = True
    write_title(ws, "COST FACTORS — Plast Fab Works Pvt. Ltd.", row=1)

    ws.row_dimensions[2].height = 30
    ws.merge_cells("A2:E2")
    ci = ws["A2"]
    ci.value = "Enter each value as a decimal fraction of annual revenue (e.g. 0.13 = 13%)"
    ci.font = Font(name="Calibri", size=10, bold=True, color=C_NAVY)
    ci.fill = fill(C_YELLOW); ci.alignment = center(); ci.border = thin_border()

    write_col_headers(ws, row=3)

    CAT_LABELS = {"operational": "Operational Costs", "strategic": "Strategic Investment", "financial": "Financial"}
    prev_cat = None
    row = 4
    for i, (key, label, cat) in enumerate(COST_FACTORS, start=1):
        if cat != prev_cat:
            section_header(ws, row, f"  {CAT_LABELS[cat]}", bg=C_ORANGE)
            row += 1; prev_cat = cat
        c = data_row(ws, row=row, num=i, label=label,
                     note="Decimal fraction (e.g. 0.13)", key=key,
                     value=COST_DATA.get(key, 0))
        c.number_format = "0.0000"
        row += 1

    ws.freeze_panes = "A4"


def build_kpi_sheet(wb):
    ws = wb.create_sheet("KPI Factors")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 6, "B": 50, "C": 18, "D": 40, "E": 0.1})
    ws.column_dimensions["E"].hidden = True
    write_title(ws, "KPI FACTORS — Plast Fab Works Pvt. Ltd.", row=1)

    ws.row_dimensions[2].height = 28
    ws.merge_cells("A2:E2")
    ci = ws["A2"]
    ci.value = "Select Yes if your organisation actively tracks this KPI, No if not."
    ci.font = Font(name="Calibri", size=10, bold=True, color=C_NAVY)
    ci.fill = fill(C_YELLOW); ci.alignment = center(); ci.border = thin_border()

    write_col_headers(ws, row=3)

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', showDropDown=False, showErrorMessage=True)
    ws.add_data_validation(dv_yn)

    CAT_LABELS = {
        "operational": "Operational Excellence", "quality": "Quality & Safety",
        "performance": "Performance Metrics", "financial": "Financial Performance",
        "customer": "Customer & People",
    }
    prev_cat = None
    row = 4
    for i, (key, label, cat) in enumerate(KPI_FACTORS, start=1):
        if cat != prev_cat:
            section_header(ws, row, f"  {CAT_LABELS[cat]}", bg=C_ORANGE)
            row += 1; prev_cat = cat
        data_row(ws, row=row, num=i, label=label, note="Select: Yes or No",
                 key=key, value=KPI_DATA.get(key, "No"), dv=dv_yn)
        row += 1

    ws.freeze_panes = "A4"


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_company_sheet(wb)
    build_barrier_sheet(wb)
    build_cost_sheet(wb)
    build_kpi_sheet(wb)

    wb.save(OUTPUT)
    print(f"Template saved: {OUTPUT}")


if __name__ == "__main__":
    main()
